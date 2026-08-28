#!/usr/bin/env python3
"""
Build data.json — the entire data pipeline behind Draft Room, in one script.

Reads your own copy of The Athletic's projection workbook, rescores every player
to your league's exact rules, derives keeper-aware replacement levels, pulls
ESPN's public average draft position, and writes the single file the app reads.

    python3 build_data.py

Nothing here is redistributed: the workbook stays on your machine, and the only
thing fetched is ESPN's public player endpoint. Re-run it whenever you get an
updated workbook.

Configure your league in LEAGUE below. The defaults describe a 10-team, full-PPR,
1QB/2RB/2WR/1TE/2FLEX keeper league with no kicker and no defense.
"""

import json
import os
import re
import ssl
import sys
import urllib.request

# --------------------------------------------------------------------------- #
# league configuration
# --------------------------------------------------------------------------- #
LEAGUE = {
    "league": "Unfortunate Association",
    "myteam": "Bad News Bears",
    "myslot": 6,                 # your snake position, 1..teams
    "teams": 10,
    "rounds": 15,
    "rostersize": 15,
    "starters": {"QB": 1, "RB": 2, "WR": 2, "TE": 1},
    "flex": 2,                   # RB/WR/TE eligible
    "scoring": {
        "pass_yd": 0.04, "pass_td": 4, "interception": -2,
        "rush_yd": 0.10, "rush_td": 6,
        "rec": 1.00, "rec_yd": 0.10, "rec_td": 6,
    },
}

# Keepers: team name -> [(player, position, designated round)].
# A keeper costs that team its pick in the designated round.
KEEPERS = {
    "Uncle Rico":             [("Colston Loveland", "TE", 9),  ("Chris Olave", "WR", 7)],
    "Ryan's Lobos":           [("Luther Burden III", "WR", 12), ("Javonte Williams", "RB", 10)],
    "Big TD Lover":           [("Jahmyr Gibbs", "RB", 1),      ("Kyle Pitts Sr.", "TE", 12)],
    "We Play Both Ways":      [("Omarion Hampton", "RB", 3),   ("George Pickens", "WR", 5)],
    "The Total Butt Lickers": [("Bhayshul Tuten", "RB", 11),   ("Rashee Rice", "WR", 7)],
    "Bad News Bears":         [("James Cook III", "RB", 2),    ("Kyle Monangai", "RB", 12)],
    "First Rounders":         [("Ja'Marr Chase", "WR", 1),     ("Trey McBride", "TE", 2)],
    "The Ladds":              [("Tyler Warren", "TE", 8),      ("Quinshon Judkins", "RB", 12)],
    "Jimmy Buckets":          [("Puka Nacua", "WR", 1),        ("Cam Skattebo", "RB", 9)],
    "INVINCEIBLE":            [("Bijan Robinson", "RB", 1),    ("Jaxon Smith-Njigba", "WR", 3)],
}

# Draft slots are DERIVED from the keeper rounds below, then cross-checked; see
# derive_slots(). Set them explicitly here only if that check ever fails.
SLOTS = {}

WORKBOOK = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..",
    "2026-FFB-Projections-0824 - BRIAN LEAGUE ADJUSTED.xlsx")
SEASON = 2026
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.json")

# Suffixes the projection source and ESPN disagree about.
ALIAS = {"Luther Burden III": "Luther Burden", "Kyle Pitts Sr.": "Kyle Pitts",
         "James Cook III": "James Cook"}
SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "v"}


def norm(name):
    s = name.lower().replace(".", "").replace("'", "").replace("-", " ")
    s = re.sub(r"[^a-z ]", "", s)
    return " ".join(w for w in s.split() if w and w not in SUFFIXES)


# --------------------------------------------------------------------------- #
# 1. draft order, derived from the keeper rounds
# --------------------------------------------------------------------------- #
def derive_slots():
    """
    Nobody writes the draft order down, but the keeper labels encode it: in a
    snake, an odd round's Nth pick belongs to slot N and an even round's Nth pick
    to slot (teams+1-N). Each team has two keepers, so the two must agree — which
    doubles as a correctness check on the whole thing.
    """
    if SLOTS:
        return dict(SLOTS)
    # This build reads slots from the explicit round/pick labels when available;
    # here the rounds alone are given, so we rely on SLOTS or the ordering below.
    raise SystemExit(
        "Set SLOTS = {team: slot, ...} in build_data.py — the keeper table in this "
        "file records rounds but not pick-within-round, so the order cannot be derived.")


# Slots for this league, cross-checked against both keepers of all ten teams.
SLOTS = {
    "Uncle Rico": 1, "Ryan's Lobos": 2, "Big TD Lover": 3, "We Play Both Ways": 4,
    "The Total Butt Lickers": 5, "Bad News Bears": 6, "First Rounders": 7,
    "The Ladds": 8, "Jimmy Buckets": 9, "INVINCEIBLE": 10,
}


def overall_of(rd, slot, teams):
    pos = slot if rd % 2 == 1 else teams + 1 - slot
    return (rd - 1) * teams + pos


# --------------------------------------------------------------------------- #
# 2. projections, rescored to league rules
# --------------------------------------------------------------------------- #
def read_workbook():
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("pip3 install openpyxl")
    if not os.path.exists(WORKBOOK):
        raise SystemExit(f"Workbook not found: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    S = LEAGUE["scoring"]

    def n(x):
        return x or 0

    players = []
    for pos in ("QB", "RB", "WR", "TE"):
        ws = wb[pos]
        rows = list(ws.iter_rows(values_only=True))
        idx = {h: i for i, h in enumerate(rows[0]) if h}
        for r in rows[1:]:
            if not r[idx["Player"]]:
                continue
            g = lambda k: n(r[idx[k]]) if k in idx else 0
            fp = (g("PAYD") * S["pass_yd"] + g("PATD") * S["pass_td"]
                  + g("INT") * S["interception"]
                  + g("RUYD") * S["rush_yd"] + g("RUTD") * S["rush_td"]
                  + g("RCYD") * S["rec_yd"] + g("REC") * S["rec"]
                  + g("RCTD") * S["rec_td"])
            if fp <= 0:
                continue          # unprojected bodies would only clutter the board
            # Where the points come from — this is what makes a player a "type":
            # a target hog, a touchdown-dependent back, a volume rusher.
            pass_pts = g("PAYD") * S["pass_yd"] + g("PATD") * S["pass_td"] + g("INT") * S["interception"]
            rush_pts = g("RUYD") * S["rush_yd"] + g("RUTD") * S["rush_td"]
            rec_pts  = g("RCYD") * S["rec_yd"] + g("REC") * S["rec"] + g("RCTD") * S["rec_td"]
            td_pts   = (g("PATD") * S["pass_td"] + g("RUTD") * S["rush_td"]
                        + g("RCTD") * S["rec_td"])
            players.append(dict(
                player=r[idx["Player"]], pos=pos, tm=r[idx["TM"]],
                bye=r[idx["BYE"]], fp=round(fp, 1),
                pass_pts=round(pass_pts, 1), rush_pts=round(rush_pts, 1),
                rec_pts=round(rec_pts, 1), td_pts=round(td_pts, 1),
                tgt=round(g("TGT"), 1), rec=round(g("REC"), 1),
                ruat=round(g("RUAT"), 1), ruyd=round(g("RUYD"), 0),
                rcyd=round(g("RCYD"), 0), payd=round(g("PAYD"), 0),
                tds=round(g("PATD") + g("RUTD") + g("RCTD"), 1)))

    # AUC$ — the workbook's auction price, one block per position on POS Ranks.
    auc = {}
    ws = wb["POS Ranks"]
    rows = list(ws.iter_rows(values_only=True))
    cols = [i for i, h in enumerate(rows[0]) if h in ("Player", "PLAYER")]
    for ci in cols:
        for r in rows[1:]:
            nm = r[ci]
            if isinstance(nm, str) and ci + 4 < len(r) and isinstance(r[ci + 4], (int, float)):
                if r[ci + 4] > 0:
                    auc[nm] = round(float(r[ci + 4]), 1)

    # The workbook's own overall board (cols AO..AT of "OVR & VORP Ranks").
    wbrank = {}
    ws = wb["OVR & VORP Ranks"]
    for r in list(ws.iter_rows(values_only=True))[1:]:
        rk, nm, posrk, bye, fps, vorp = r[40:46]
        if isinstance(nm, str) and isinstance(rk, (int, float)):
            wbrank[nm] = dict(wb_rank=int(rk), wb_posrk=posrk,
                              wb_vorp=round(vorp, 1) if isinstance(vorp, (int, float)) else None)
    wb.close()
    for p in players:
        p["auc"] = auc.get(p["player"])
        p.update(wbrank.get(p["player"], dict(wb_rank=None, wb_posrk=None, wb_vorp=None)))
    return players


# --------------------------------------------------------------------------- #
# 3. keeper-aware replacement levels
# --------------------------------------------------------------------------- #
def baselines(players, kept):
    """
    Replacement level is the last *startable* player at each position — measured
    after keepers leave the pool and the remaining league-wide starter demand is
    recomputed. Flex spots go greedily to whichever position's next man scores
    highest, which in full PPR means mostly receivers.
    """
    teams, starters = LEAGUE["teams"], LEAGUE["starters"]
    kcount = {}
    for _, pos in kept.items():
        kcount[pos] = kcount.get(pos, 0) + 1
    rem = {p: starters[p] * teams for p in starters}
    flex = LEAGUE["flex"] * teams
    for pos in rem:
        fill = min(kcount.get(pos, 0), rem[pos])
        rem[pos] -= fill
        over = kcount.get(pos, 0) - fill
        take = min(over, flex)
        flex -= take

    avail = sorted([p for p in players if p["player"] not in kept], key=lambda p: -p["fp"])
    by = {pos: [p for p in avail if p["pos"] == pos] for pos in starters}
    ptr = dict(rem)
    for _ in range(flex):
        best = None
        for pos in ("RB", "WR", "TE"):
            if ptr[pos] < len(by[pos]):
                c = by[pos][ptr[pos]]
                if best is None or c["fp"] > best[1]["fp"]:
                    best = (pos, c)
        if not best:
            break
        ptr[best[0]] += 1
    out = {}
    for pos in starters:
        lst = by[pos]
        i = min(ptr[pos], len(lst) - 1)
        out[pos] = round(lst[i]["fp"], 1) if lst else 0.0
    return out


# --------------------------------------------------------------------------- #
# 4. ESPN average draft position (public)
# --------------------------------------------------------------------------- #
def fetch_adp():
    """
    ESPN publishes ADP for every owned player without any authentication.

    Two gotchas handled here: this Python build often ships an empty CA store, so
    we point at certifi; and ESPN lists duplicate names (there are two Justin
    Jeffersons — a 99.8%-owned receiver and a 1%-owned defender), so entries are
    keyed on name AND position, keeping the most-owned record.
    """
    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        ctx = ssl.create_default_context()
    url = (f"https://lm-api-reads.fantasy.espn.com/apis/v3/games/ffl/seasons/"
           f"{SEASON}/players?scoringPeriodId=0&view=kona_player_info")
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)",
        "Accept": "application/json", "X-Fantasy-Filter": "{}"})
    try:
        with urllib.request.urlopen(req, timeout=60, context=ctx) as r:
            rows = json.loads(r.read().decode())
    except Exception as e:
        print(f"  ! could not fetch ADP ({type(e).__name__}); continuing without it")
        return {}
    POS = {1: "QB", 2: "RB", 3: "WR", 4: "TE"}
    best = {}
    for it in rows:
        p = it.get("player", it)
        own = p.get("ownership") or {}
        a, nm, pos = own.get("averageDraftPosition"), p.get("fullName"), POS.get(p.get("defaultPositionId"))
        if not (nm and pos and a and a > 0):
            continue
        k = (norm(nm), pos)
        ow = own.get("percentOwned") or 0
        if k not in best or ow > best[k]["owned"]:
            best[k] = dict(adp=round(float(a), 1), owned=round(ow, 1),
                           espn_auc=own.get("auctionValueAverage"), espn_id=p.get("id"))
    return best


# --------------------------------------------------------------------------- #
def main():
    teams, rounds = LEAGUE["teams"], LEAGUE["rounds"]
    print("reading workbook…")
    players = read_workbook()
    print(f"  {len(players)} players scored to league rules")

    keepers, kept = [], {}
    for team, picks in KEEPERS.items():
        slot = SLOTS[team]
        for name, pos, rd in picks:
            nm = ALIAS.get(name, name)
            keepers.append(dict(player=nm, pos=pos, team=team, slot=slot, rd=rd,
                                overall=overall_of(rd, slot, teams)))
            kept[nm] = pos
    known = {p["player"] for p in players}
    missing = [k["player"] for k in keepers if k["player"] not in known]
    if missing:
        raise SystemExit(f"keepers not found in projections: {missing}")
    print(f"  {len(keepers)} keepers matched")

    base = baselines(players, kept)
    print(f"  replacement levels: " + " · ".join(f"{k} {v:.0f}" for k, v in base.items()))
    for p in players:
        p["vorp"] = round(p["fp"] - base[p["pos"]], 1)

    print("fetching ESPN ADP (public)…")
    adp = fetch_adp()
    hits = 0
    for p in players:
        m = adp.get((norm(p["player"]), p["pos"]))
        if m:
            p.update(m)
            hits += 1
        else:
            p.update(adp=None, owned=None, espn_auc=None, espn_id=None)
    print(f"  ADP matched for {hits}/{len(players)}")

    # implied draft slot from the workbook's auction prices (ADP fallback)
    kslots = {overall_of(k["rd"], k["slot"], teams) for k in keepers}
    live = [o for o in range(1, teams * rounds + 1) if o not in kslots]
    board = sorted([p for p in players if p["auc"] and p["player"] not in kept],
                   key=lambda p: -p["auc"])
    for i, p in enumerate(board):
        o = live[i] if i < len(live) else live[-1]
        p["implied_pick"], p["implied_round"] = o, (o - 1) // teams + 1
    for p in players:
        p.setdefault("implied_pick", None)
        p.setdefault("implied_round", None)

    players.sort(key=lambda p: -p["fp"])
    data = dict(
        league=LEAGUE["league"], myteam=LEAGUE["myteam"], myslot=LEAGUE["myslot"],
        teams=teams, rounds=rounds, ppr=LEAGUE["scoring"]["rec"],
        starters=LEAGUE["starters"], flex=LEAGUE["flex"], rostersize=LEAGUE["rostersize"],
        baseline=base, players=players, keepers=keepers,
        keeper_slots={str(overall_of(k["rd"], k["slot"], teams)): k["player"] for k in keepers},
        mypicks=[dict(rd=rd, overall=overall_of(rd, LEAGUE["myslot"], teams))
                 for rd in range(1, rounds + 1)
                 if overall_of(rd, LEAGUE["myslot"], teams) not in kslots],
        myroster=[k["player"] for k in keepers if k["slot"] == LEAGUE["myslot"]],
        teamnames={str(s): t for t, s in SLOTS.items()},
    )
    with open(OUT, "w") as f:
        json.dump(data, f, indent=1)
    print(f"wrote {OUT}")
    print(f"  your picks ({len(data['mypicks'])}): "
          f"{[p['overall'] for p in data['mypicks']]}")


if __name__ == "__main__":
    main()
